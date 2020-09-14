import openpyxl,requests
import smtplib
from email.mime.text import MIMEText#支持html格式
from email.mime.multipart import MIMEMultipart
from email.header import Header
from log.log import Logger

class LoadExcel:
    def __init__(self,path):
        self.path = path
        self.excel = openpyxl.load_workbook(path)
        self.sheet = self.excel.active#当前打开的sheet

    def get_sheet_by_index(self,index):
        return self.excel.sheetnames[index]

    def set_sheet_by_index(self,index):
        self.sheet = self.excel.sheetnames[index]

    def get_cell_value(self,row,col):
        return self.sheet.cell(row,col).value

    def set_cell_value(self,row,col,value):
        self.sheet.cell(row,col,value=value)
    def get_max_row(self):
        return self.sheet.max_row

    def save_excel(self, path):
        self.excel.save(path)
    def close_excel(self):
        self.excel.close()

#发送html格式邮件（需要修改报告源码）
def send_mail(email_path):
    from config.config import host, port, password, sender, receiver
    with open(email_path, 'rb') as f:
        content = f.read()
    host = host
    port = port
    sender = sender
    password = password
    receiver = receiver
    message = MIMEText(content, "HTML", "UTF-8")
    message["Subject"] = "接口测试"
    message["From"] = sender
    message["To"] = receiver
    try:
        smtp = smtplib.SMTP(host, port)
        smtp.login(sender,password)
        smtp.sendmail(sender, receiver, message.as_string())
        Logger().log().info("发送邮件成功")
    except smtplib.SMTPException as e:
        Logger().log().info("发送邮件失败，失败信息：{}".format(e))

#发送附件邮件
def send_email(email_path):
    from config.config import host, port, password, sender, receiver
    message = MIMEMultipart()
    #邮件内容
    text = """
    请输入你想说的邮件内容
    """
    message.attach(MIMEText(_text=text, _subtype='plain', _charset="utf-8"))
    #需要发送的附件的路径
    with open(email_path, 'rb') as f:
        content = f.read()
    att1 = MIMEText(content, "base64", "utf-8")
    att1["Content-Type"] = 'application/octet-stream'
    att1['Content-Disposition'] = 'attachment; filename = "report.xlsx"'#发送相应格式html或者xlsx
    message.attach(att1)

    #邮件主题
    message["Subject"] = Header("柠檬web商城", "utf-8").encode()
    message["From"] = Header("tianmeng", "utf-8")
    message["To"] = Header('tianmeng_wxk', "utf-8")

    try:
        smtp = smtplib.SMTP()
        #smtp = smtplib.SMTP_SSL('smtp.163.com', 465)
        smtp.connect(host=host, port=port)
        smtp.login(user=sender, password=password)
        sender = sender
        receiver = [receiver]
        smtp.sendmail(sender, receiver, message.as_string())
        Logger().log().info("发送邮件成功")
        return email_path
    except smtplib.SMTPException as e:
        Logger().log().info("发送邮件失败，失败信息：{}".format(e))



if __name__ == '__main__':
    path = '../cases/login.xlsx'
    excel = LoadExcel(path)
    print(excel.get_cell_value(2,2))