from key_word.key_word import *
from common.common import LoadExcel
from openpyxl.styles import PatternFill,Font
from log.log import Logger
from config.config import case
import unittest,os,sys,time
path = 'E:\script\web_keyword_lemon'
sys.path.append(path)

path = case
operator_id = 1
operator_name = 2
keyword = 3
loc_type = 4
loc_ex = 5
operator_value = 6
expect = 7
result = 8

def test_login():
    excel = LoadExcel(path)
    for rownum in range(2,excel.get_max_row()+1):
        operator_name1 = excel.get_cell_value(rownum,operator_name)

        keyword1 = excel.get_cell_value(rownum,keyword)
        loc_type1 = excel.get_cell_value(rownum, loc_type)
        loc_ex1 = excel.get_cell_value(rownum, loc_ex)
        operator_value1 = excel.get_cell_value(rownum, operator_value)
        expect1 = excel.get_cell_value(rownum,expect)
        if keyword1 == 'ivercode':
            func = keyword1+"()"
            func = eval(func)

            Logger().log().debug("操作描述：{}".format(operator_name1))

            # sheet.cell(row, 9).value = "pass"
            Logger().log().debug("func值为：{}".format(func))
            excel.set_cell_value(rownum+1,operator_value,func)


        elif "assert" in keyword1:
            func = '{}("{}","{}","{}")'.format(keyword1, loc_type1, loc_ex1, expect1)
            Logger().log().debug("操作描述：{}".format(operator_name1))
            status = eval(func)
            if status == True:

                # excel.set_cell_value(rownum, result, "pass")
                # excel.get_cell_value(rownum,result).fill = PatternFill('solid', fgColor='66ff00')
                # excel.get_cell_value(rownum, result).font = Font(bold=True)
                excel.sheet.cell(rownum, result).value = "pass"
                excel.sheet.cell(rownum, result).fill = PatternFill('solid', fgColor='66ff00')
                excel.sheet.cell(rownum, result).font = Font(bold=True)
            else:
                # excel.set_cell_value(rownum, result, "false")
                # excel.get_cell_value(rownum,result).fill = PatternFill('solid', fgColor='FF0000')
                # excel.get_cell_value(rownum, result).font = Font(bold=True)
                excel.sheet.cell(rownum, result).value = "false"
                excel.sheet.cell(rownum, result).fill = PatternFill('solid', fgColor='FF0000')
                excel.sheet.cell(rownum, result).font = Font(bold=True)
        else:
            if loc_type1==loc_ex1==operator_value1==None:
                func = keyword1+"()"
                Logger().log().debug("操作描述：{}".format(operator_name1))
            elif operator_value1==None:
                func = '{}("{}","{}")'.format(keyword1,loc_type1,loc_ex1)
                Logger().log().debug("操作描述：{}".format(operator_name1))
            elif loc_type1==loc_ex1==None:
                func = '{}("{}")'.format(keyword1,operator_value1)
                Logger().log().debug("操作描述：{}".format(operator_name1))
            else:
                func = '{}("{}","{}","{}")'.format(keyword1,loc_type1,loc_ex1,operator_value1)
                Logger().log().debug("操作描述：{}".format(operator_name1))
            eval(func)
    excel.save_excel(path)
if __name__ == '__main__':
    test_login()
